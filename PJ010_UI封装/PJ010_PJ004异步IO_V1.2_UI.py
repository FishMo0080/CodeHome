import asyncio
import aiohttp
import aiofiles
from openpyxl import load_workbook, Workbook
from datetime import datetime
import warnings
import tkinter as tk
from tkinter import messagebox
import openpyxl
import glob
def run_program():
    # 获取输入框中的内容
    cinput = entry.get()
    # 在这里编写程序的运行逻辑
    warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')  #主动关闭表格样式的告警
    datet = datetime.now().strftime('%Y-%m-%d')                                 # 获取当前日期以便用时间命名文件夹
    new_title = 'PMIS任务数据汇总' + datet
    async def aiodownload(proj_id):
        payload = dict(projectCode=proj_id)
        async with aiohttp.ClientSession() as session:
            async with session.post(url,data=payload,headers=headers) as resp:
                if resp.status == 200:            
                    result = await resp.read()
                    async with aiofiles.open(f"{proj_id}.xlsx", 'wb') as f:
                        await f.write(result)

    async def main():
            ready_wb = load_workbook("ready/ready.xlsx")
            ready_ws = ready_wb.active
            proj_ids = [cell.value for cell in ready_ws["A"]]
            tasks = [asyncio.create_task(aiodownload(proj_id)) for proj_id in proj_ids]
            await asyncio.wait(tasks)

    if __name__ == '__main__':
        headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "User-Agent": "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729)",
        "Cookie": f"pmis.session.id={cinput}",
        }
        url = "http://pmis.iim.gmcc.net/a/portal/overview/exportProjectTasks" 
        asyncio.run(main())

        # 以下将得到的Excel表汇总在一个Excel文件
        workbook = openpyxl.Workbook()
        sheet_total = workbook.active
        filenames = glob.glob('*.xlsx')        
        for filename in filenames:            
            workbook_temp = openpyxl.load_workbook(filename)    
            sheet = workbook_temp.active    
            for row in sheet.iter_rows(min_row=3, min_col=1,max_col=sheet.max_column,max_row=sheet.max_row):
                row_data = [col.value for col in row]
                sheet_total.append(row_data)
            # break
        workbook.save(new_title + '.xlsx')
        # 弹出对话框，提示用户数据已导出
        messagebox.showinfo("提示", "你的数据已导出，并汇总。")
        # 关闭主窗口
        root.destroy()
    else:
        # 清空参数输入框
        entry.delete(0, tk.END)
        # 弹出错误提示框
        tk.messagebox.showerror("注意！", "ID错误或过期，请重新获取！")
# 创建主窗口
root = tk.Tk()
root.title("PMIS任务数据导出汇总——by Fish")
# 创建标签和输入框
label = tk.Label(root, text="请输入刚刚获取的pmis.session.id：")
label.pack(pady=5)
entry = tk.Entry(root, width=40)
entry.pack(pady=5)
# 创建“运行程序”按钮
button = tk.Button(root, text="执行导出汇总", command=run_program)
button.pack(pady=5)
# 创建提示语
tip1 = tk.Label(root, text="**请耐心等待一下。", justify="left",highlightbackground="yellow")
tip2 = tk.Label(root, text="**执行时间根据项目数量及任务内容而不同", justify="left")
tip2.pack(side="bottom")
tip1.pack(side="bottom")
# 进入消息循环
root.mainloop()