import os
import glob
import openpyxl
import requests
from openpyxl import load_workbook, Workbook     
from datetime import datetime
import warnings
# 以上导入第三方库
import tkinter as tk
from tkinter import messagebox
def run_program():
    # 获取输入框中的内容
    cinput = entry.get()
    # 在这里编写程序的运行逻辑
    warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')  #主动关闭表格样式的告警
    datet = datetime.now().strftime('%Y-%m-%d')                                 # 获取当前日期以便用时间命名文件夹
    new_title = 'PMIS任务数据汇总' + datet
    url = ' http://pmis.iim.gmcc.net/a/portal/overview/exportProjectTasks'      # 目标网址
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "User-Agent": "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729)",
        "Cookie": f"pmis.session.id={cinput}",
        }
    # 以上编写请求头字典
    # 以下读取ready.xlsx文件的A列数据
    payload = dict(projectCode="B21302110AA00") #参数字典因为只有1个用这个编写方式。
    r = requests.post(url=url, data=payload,headers=headers)
    # 以上进行一次post，判断返回的响应类型
    if r.headers.get('Content-Type') == "application/octet-stream;charset=utf-8":
        ready_wb = load_workbook("ready/ready.xlsx")
        ready_ws = ready_wb.active
        proj_ids = [cell.value for cell in ready_ws["A"]]
        for proj_id in proj_ids:
            payload = dict(projectCode=proj_id) #参数字典因为只有1个用这个编写方式。
            r = requests.post(url=url, data=payload,headers=headers)
            if r.status_code == 200: #删除本段判断代码将会影响表格汇总，未知原因
            # 打开一个新文件，将文件流写入其中
                with open(f"{proj_id}.xlsx", 'wb') as f:
                    f.write(r.content)            
        r.close
        # 以下将得到的Excel表汇总在一个Excel文件
        workbook = openpyxl.Workbook()
        sheet_total = workbook.active
        filenames = glob.glob('*.xlsx')        
        for filename in filenames:            
            workbook_temp = openpyxl.load_workbook(filename)    
            sheet = workbook_temp.active    
            for row in sheet.iter_rows(min_row=3, min_col=1,max_col=sheet.max_column,max_row=sheet.max_row):
                row_data = [col.value for col in row]
                sheet_total.append(row_data)
            # break
        workbook.save(new_title + '.xlsx')
        # 弹出对话框，提示用户数据已导出
        messagebox.showinfo("提示", "你的数据已导出，并汇总。")
        # 关闭主窗口
        root.destroy()
    else:
        # 清空参数输入框
        entry.delete(0, tk.END)
        # 弹出错误提示框
        tk.messagebox.showerror("注意！", "ID错误或过期，请重新获取！")
# 创建主窗口
root = tk.Tk()
root.title("PMIS任务数据导出汇总——by Fish")
# 创建标签和输入框
label = tk.Label(root, text="请输入刚刚获取的pmis.session.id：")
label.pack(pady=5)
entry = tk.Entry(root, width=40)
entry.pack(pady=5)
# 创建“运行程序”按钮
button = tk.Button(root, text="执行导出汇总", command=run_program)
button.pack(pady=5)
# 创建提示语
tip1 = tk.Label(root, text="**请耐心等待一下。", justify="left",highlightbackground="yellow")
tip2 = tk.Label(root, text="**执行时间根据项目数量及任务内容而不同", justify="left")
tip2.pack(side="bottom")
tip1.pack(side="bottom")
# 进入消息循环
root.mainloop()