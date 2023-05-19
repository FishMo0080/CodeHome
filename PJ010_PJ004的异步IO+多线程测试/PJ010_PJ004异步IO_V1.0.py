import os
import glob
import openpyxl
import requests
from openpyxl import load_workbook, Workbook     
import asyncio
import aiohttp
import aiofiles
# 以上导入第三方库
headers = {
    "Content-Type": "application/x-www-form-urlencoded",
    "User-Agent": "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729)",
    "Cookie": "pmis.session.id=b88a18e2b54b4ce8907df238cf187122",
    }
async def fetch(session, url):
    async with session.post(url) as response:
        return await response.content()
    
async def main():
    async with aiohttp.ClientSession() as session:
        ready_wb = load_workbook("ready/ready.xlsx")
        ready_ws = ready_wb.active
        proj_ids = [cell.value for cell in ready_ws["A"]]
        urls = [
            f"http://pmis.iim.gmcc.net/a/portal/overview/exportProjectTasks?projectCode={proj_id}"
            for proj_id in proj_ids
        ]
        tasks = [asyncio.create_task(fetch(session, url)) for url in urls]
        htmls = await asyncio.gather(*tasks)
        print(htmls)

if __name__ == '__main__':
    asyncio.run(main())



'''
r = requests.post(url=urls[0], headers=headers)
if r.status_code == 200:
# 打开一个新文件，将文件流写入其中
    with open(f"{proj_ids[0]}.xlsx", 'wb') as f:
        f.write(r.content)   
r.close
'''

'''
url = ' http://pmis.iim.gmcc.net/a/portal/overview/exportProjectTasks'      # 目标网址
headers = {
    "Content-Type": "application/x-www-form-urlencoded",
    "User-Agent": "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729)",
    "Cookie": "pmis.session.id=6d45841ac9984b38b129ffe9cbacca00",
    }
# 以上编写请求头字典
# 以下读取ready.xlsx文件的A列数据
ready_wb = load_workbook("ready/ready.xlsx")
ready_ws = ready_wb.active
proj_ids = [cell.value for cell in ready_ws["A"]]
for proj_id in proj_ids:
    payload = dict(projectCode=proj_id) #参数字典因为只有1个用这个编写方式。
    r = requests.post(url=url, data=payload,headers=headers)
    if r.status_code == 200:
    # 打开一个新文件，将文件流写入其中
        with open(f"{proj_id}.xlsx", 'wb') as f:
            f.write(r.content)            
r.close
'''