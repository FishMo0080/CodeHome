import openpyxl

# 打开Excel文件
wb = openpyxl.load_workbook('U-learning（1~4月）V4.xlsx')
# 选择工作表
ws = wb.active

# 计算行数和列数
rows = ws.max_row
cols = ws.max_column

# 新建一个空的工作表
new_ws = wb.create_sheet('Transposed')

# 初始行数和列数为0
new_row = 1
new_col = 1

# 遍历原来的工作表
for row in range(1, rows+1):
    for col in range(1, cols+1):
        # 根据当前行数和列数计算新的行数和列数
        new_row = (row-1) // 6 + 1
        new_col = (col-1) % 6 + 1
        # 将转置后的值写入到新的工作表中
        new_ws.cell(row=new_row, column=new_col, value=ws.cell(row=row, column=col).value)

# 保存Excel文件
wb.save('U-learning（1~4月）V4-new.xlsx')