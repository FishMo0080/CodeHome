import openpyxl

#打开Excel文件
wb = openpyxl.load_workbook('U-learning（1~4月）V4.xlsx')

#选择工作表
ws = wb['Sheet1']

#获取A列单元格内容
a_column = [cell.value for cell in ws['A']]

#定义转换后的二维列表
transposed_list = []

#每6个单元格为一行，进行转置操作
for i in range(0, len(a_column), 6):
    row = a_column[i:i+6] 
    transposed_list.append(row)

#将转换后的数据写入Excel表格中
for i in range(len(transposed_list)): 
    for j in range(len(transposed_list[i])): 
        cell = ws.cell(row=j+1, column=i+1) 
        cell.value = transposed_list[i][j]

#保存Excel文件
wb.save('U-learning（1~4月）V4-new.xlsx')