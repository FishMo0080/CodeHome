import os
import glob
import openpyxl
from openpyxl import load_workbook, Workbook
# 读取ready.xlsx文件的A列数据
ready_wb = load_workbook("ready/ready.xlsx")
ready_ws = ready_wb.active
proj_ids = [cell.value for cell in ready_ws["A"]]
# 遍历文件夹中的Excel文件，按项目编号写入新的Excel文件
for proj_id in proj_ids:
    # 拼接Excel文件的路径
    file_path = os.path.join("G:\RXL", f"{proj_id}.xlsx")
    if not os.path.exists(file_path):
        # 如果文件不存在，跳过这个项目
        continue
    # 读取Excel文件
    wb = load_workbook(file_path)
    ws = wb.active
    # 创建新的Excel文件，以项目编号命名
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = proj_id
    # 遍历Excel中的每一个单元格，将数据写入新的Excel文件
    for row in ws.iter_rows():
        new_ws.append([cell.value for cell in row])
    # 保存新的Excel文件
    new_wb.save(f"{proj_id}.xlsx")
# 将得到的数据汇总在一个Excel文件
workbook = openpyxl.Workbook()
sheet_total = workbook.active
filenames = glob.glob('*.xlsx')
print(len(filenames))
for filename in filenames:
    print(filename)
    workbook_temp = openpyxl.load_workbook(filename)
    sheet = workbook_temp.active
    for row in sheet.iter_rows(min_row=2, min_col=1,max_col=sheet.max_column,max_row=sheet.max_row):
        row_data = [col.value for col in row]

        sheet_total.append(row_data)
    # break
workbook.save('汇总.xlsx')