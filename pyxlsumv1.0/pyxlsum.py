import glob
import openpyxl
workbook = openpyxl.Workbook()
sheet_total = workbook.active
filenames = glob.glob('*.xlsx')
print(len(filenames))

for filename in filenames:
    print(filename)
    workbook_temp = openpyxl.load_workbook(filename)
    sheet = workbook_temp.active
    for row in sheet.iter_rows(min_row=2, min_col=1,max_col=sheet.max_column,max_row=sheet.max_row):
        row_data = [col.value for col in row]

        sheet_total.append(row_data)
    # break
workbook.save('汇总.xlsx')