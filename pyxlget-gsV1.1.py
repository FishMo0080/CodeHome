import glob
import os
import openpyxl
files = glob.glob('D:\\Ptest\\*.xlsx')
print(len(files))
result_wb = openpyxl.Workbook()
# 遍历文件夹
for root, dirs, files in os.walk("D:\\Ptest"):
    # 遍历文件
    for file in files:
        print(file)
        # 判断文件是否为Excel文件
        if file.endswith(".xlsx"):
            # 打开Excel文件
            wb = openpyxl.load_workbook(os.path.join(root, file))
            # 获取sheet1
            ws = wb["表二"]
            # 获取a1和a2的数据
            a1 = ws["A1"].value
            a2 = ws["D6"].value
            # 将数据汇总到新建的表格中
            result_ws = result_wb.active
            result_ws.append([file, a1, a2])
            result_wb.save("result.xlsx")