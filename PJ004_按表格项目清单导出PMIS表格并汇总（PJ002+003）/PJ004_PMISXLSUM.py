import os
import glob
import openpyxl
import requests
from openpyxl import load_workbook, Workbook     
from datetime import datetime
import warnings
# 以上导入第三方库
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')  #主动关闭表格样式的告警
datet = datetime.now().strftime('%Y-%m-%d')                                 # 获取当前日期以便用时间命名文件夹
new_title = 'PMIS任务数据汇总' + datet
url = ' http://pmis.iim.gmcc.net/a/portal/overview/exportProjectTasks'      # 目标网址
headers = {
    "Content-Type": "application/x-www-form-urlencoded",
    "User-Agent": "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729)",
    "Cookie": "pmis.session.id=9db1933d54bc49beadbff5ce79b5eabd",
    }
# 以上编写请求头字典
# 以下读取ready.xlsx文件的A列数据
ready_wb = load_workbook("ready/ready.xlsx")
ready_ws = ready_wb.active
proj_ids = [cell.value for cell in ready_ws["A"]]
for proj_id in proj_ids:
    payload = dict(projectCode=proj_id) #参数字典因为只有1个用这个编写方式。
    r = requests.post(url=url, data=payload,headers=headers)
    if r.status_code == 200:
    # 打开一个新文件，将文件流写入其中
        with open(f"{proj_id}.xlsx", 'wb') as f:
            f.write(r.content)            
#以上直接将内容写入Excel，如果大型文件可以用一下方法，1kb读取循环安全写入。
"""
            for chunk in r.iter_content(1024):
                f.write(chunk)
                r.close
"""
r.close  #关闭网页请求，如果不关闭有可能会由于请求太多端口问题导致暂时无法请求（这要思考一下要在循环里面关，还是在循环外面关）
# 以下将得到的数据汇总在一个Excel文件
workbook = openpyxl.Workbook()
sheet_total = workbook.active
filenames = glob.glob('*.xlsx')
print(len(filenames))
for filename in filenames:
    print(filename)
    workbook_temp = openpyxl.load_workbook(filename)    
    sheet = workbook_temp.active    
    for row in sheet.iter_rows(min_row=3, min_col=1,max_col=sheet.max_column,max_row=sheet.max_row):
        row_data = [col.value for col in row]
        sheet_total.append(row_data)
    # break
workbook.save(new_title + '.xlsx')
